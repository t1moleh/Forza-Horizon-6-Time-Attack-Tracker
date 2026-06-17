"""Schoen formatierter Excel-Export (openpyxl) aus lap_times.csv.

Blaetter:
  * Bestzeiten   - beste Zeit je (Strecke, Auto), je Strecke gerankt
  * Pro Fahrzeug - alle Zeiten je Fahrzeug
  * Alle Runden  - vollstaendiges Log (chronologisch)

    py -m fh6tracker.excel_export
    py -m fh6tracker.excel_export --out FH6_Rundenzeiten.xlsx
"""
from __future__ import annotations

import argparse
import csv
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .snapshot import build_by_car, build_overall

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TRACK_FILL = PatternFill("solid", fgColor="374151")
TRACK_FONT = Font(bold=True, color="FFFFFF")
BEST_FILL = PatternFill("solid", fgColor="DCFCE7")     # gruen: Bestzeit je Strecke
CENTER = Alignment(horizontal="center")
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    ws.freeze_panes = "A2"


def _autofit(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _sheet_overall(wb: Workbook, log_path: str) -> None:
    ws = wb.active
    ws.title = "Bestzeiten"
    cols = ["Strecke", "Rang", "Auto", "Klasse", "PI", "Antrieb", "Bestzeit"]
    ws.append(cols)
    for e in build_overall(log_path):
        row = [e["track"], e["rank"], e["car_name"], e["class"], e["pi"],
               e["drivetrain"], e["time"]]
        ws.append(row)
        r = ws.max_row
        for c in range(1, len(cols) + 1):
            ws.cell(row=r, column=c).border = BORDER
        if e["rank"] == 1:
            for c in range(1, len(cols) + 1):
                ws.cell(row=r, column=c).fill = BEST_FILL
    _style_header(ws, len(cols))
    _autofit(ws, [34, 6, 30, 8, 6, 9, 12])


def _sheet_by_car(wb: Workbook, log_path: str) -> None:
    ws = wb.create_sheet("Pro Fahrzeug")
    cols = ["Fahrzeug", "Klasse", "PI", "Strecke", "Zeit", "Datum"]
    ws.append(cols)
    for car in build_by_car(log_path):
        # Trennzeile je Fahrzeug
        ws.append([f"{car['name']}  (Klasse {car['class']}, PI {car['pi']}, "
                   f"{car['lap_count']} Runden, beste {car['best']})"])
        r = ws.max_row
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(cols))
        cell = ws.cell(row=r, column=1)
        cell.fill = TRACK_FILL
        cell.font = TRACK_FONT
        for lp in car["laps"]:
            ws.append(["", car["class"], car["pi"], lp["track"], lp["time"],
                       lp["timestamp"]])
            rr = ws.max_row
            for c in range(1, len(cols) + 1):
                ws.cell(row=rr, column=c).border = BORDER
    _style_header(ws, len(cols))
    _autofit(ws, [30, 8, 6, 34, 12, 20])


def _sheet_all_laps(wb: Workbook, log_path: str) -> None:
    ws = wb.create_sheet("Alle Runden")
    cols = ["Datum/Uhrzeit", "Auto", "Klasse", "PI", "Antrieb", "Strecke", "Zeit"]
    ws.append(cols)
    if os.path.exists(log_path):
        with open(log_path, newline="", encoding="utf-8") as fh:
            for r in csv.DictReader(fh):
                ws.append([r.get("datum_uhrzeit", ""), r.get("auto", ""),
                           r.get("klasse", ""), r.get("pi", ""), r.get("antrieb", ""),
                           r.get("strecke", ""), r.get("rundenzeit", "")])
                rr = ws.max_row
                for c in range(1, len(cols) + 1):
                    ws.cell(row=rr, column=c).border = BORDER
    _style_header(ws, len(cols))
    _autofit(ws, [20, 30, 8, 6, 9, 34, 12])


def export_excel(log_path: str, out_path: str) -> str:
    wb = Workbook()
    _sheet_overall(wb, log_path)
    _sheet_by_car(wb, log_path)
    _sheet_all_laps(wb, log_path)
    wb.save(out_path)
    return out_path


def main(argv: list[str] | None = None) -> None:
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    ap = argparse.ArgumentParser(description="Excel-Export der Rundenzeiten")
    ap.add_argument("--log", default=os.path.join(root, "lap_times.csv"))
    ap.add_argument("--out", default=os.path.join(root, "FH6_Rundenzeiten.xlsx"))
    args = ap.parse_args(argv)
    path = export_excel(args.log, args.out)
    print(f"Excel geschrieben: {path}")


if __name__ == "__main__":
    main()
